"""
Create a properly formatted sample Excel file for Enjaz system.
"""

import pandas as pd
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_sample_excel():
    """Create a sample Excel file with correct structure."""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets for different subjects
    subjects = [
        ("Arabic", "03/1"),
        ("Math", "03/1"),
        ("Science", "03/2")
    ]
    
    # Sample student names
    students = [
        "أحمد محمد علي",
        "محمد أحمد سالم",
        "علي حسن محمود",
        "حسن علي أحمد",
        "سالم محمد حسن",
        "خالد أحمد علي",
        "عمر محمد سالم",
        "يوسف علي حسن"
    ]
    
    # Create dates (past dates)
    today = date.today()
    dates = [today - timedelta(days=i*7) for i in range(5, 0, -1)]  # 5 weeks ago to 1 week ago
    
    for subject, class_code in subjects:
        sheet_name = f"{class_code.replace('/', '-')} {subject}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Row 1: Headers
        ws['A1'] = "اسم الطالب"
        ws['B1'] = "Overall"
        
        # Assessment headers (starting from column H, index 7)
        for i, d in enumerate(dates):
            col_letter = chr(72 + i)  # H, I, J, K, L
            ws[f'{col_letter}1'] = f"Week {i+1}"
        
        # Row 2: Empty or additional info
        ws['A2'] = ""
        
        # Row 3: Due dates
        ws['A3'] = ""
        ws['B3'] = ""
        for i, d in enumerate(dates):
            col_letter = chr(72 + i)
            ws[f'{col_letter}3'] = d.strftime("%Y-%m-%d")
        
        # Rows 4+: Student data
        for row_idx, student_name in enumerate(students, start=4):
            ws[f'A{row_idx}'] = student_name
            ws[f'B{row_idx}'] = ""  # Overall column (will be calculated)
            
            # Assessment scores
            for i in range(len(dates)):
                col_letter = chr(72 + i)
                # Mix of scores, M, I, AB, X
                import random
                value_type = random.choice(['score', 'score', 'score', 'M', 'I', 'AB', 'X'])
                
                if value_type == 'score':
                    score = random.randint(60, 100)
                    ws[f'{col_letter}{row_idx}'] = score
                elif value_type == 'M':
                    ws[f'{col_letter}{row_idx}'] = 'M'
                elif value_type == 'I':
                    ws[f'{col_letter}{row_idx}'] = 'I'
                elif value_type == 'AB':
                    ws[f'{col_letter}{row_idx}'] = 'AB'
                else:
                    ws[f'{col_letter}{row_idx}'] = 'X'
        
        # Apply formatting
        # Header row
        header_fill = PatternFill(start_color="8A1538", end_color="8A1538", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col in range(1, 8 + len(dates)):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Date row
        date_fill = PatternFill(start_color="C9A227", end_color="C9A227", fill_type="solid")
        date_font = Font(bold=True, size=10)
        
        for i in range(len(dates)):
            col_letter = chr(72 + i)
            cell = ws[f'{col_letter}3']
            cell.fill = date_fill
            cell.font = date_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        for i in range(len(dates)):
            col_letter = chr(72 + i)
            ws.column_dimensions[col_letter].width = 12
    
    # Save file
    output_path = "/home/ubuntu/enjaz/sample_data_correct.xlsx"
    wb.save(output_path)
    print(f"✅ Sample Excel file created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_sample_excel()

