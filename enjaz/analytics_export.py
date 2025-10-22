"""
Analytics Export Module
Creates detailed analytical DataFrame with one row per student-subject combination
"""

import pandas as pd
import numpy as np
from typing import List, Dict


def get_tier(overall_pct: float) -> str:
    """
    Get tier classification based on overall percentage.
    
    Args:
        overall_pct: Overall percentage across all subjects
    
    Returns:
        str: Tier name
    """
    if overall_pct >= 90:
        return "بلاتينية"
    elif overall_pct >= 80:
        return "ذهبية"
    elif overall_pct >= 70:
        return "فضية"
    elif overall_pct >= 50:
        return "برونزية"
    elif overall_pct >= 1:
        return "يحتاج إلى تطوير"
    else:  # 0
        return "لا يستفيد من النظام"


def create_analytics_export(all_data: List[Dict]) -> pd.DataFrame:
    """
    Create analytics export DataFrame with one row per student-subject.
    
    Each row contains:
    1. student_name (string)
    2. grade (string)
    3. section (string)
    4. subject (string)
    5. subject_total_assigned (int) - total assignments for this subject
    6. subject_total_done (int) - completed assignments for this subject
    7. overall_pct_all_subjects (float, 1 decimal) - overall % across ALL subjects for this student
    8. tier (string) - classification based on overall_pct_all_subjects
    
    Args:
        all_data: List of sheet data dictionaries
    
    Returns:
        DataFrame with analytics export
    """
    
    # Step 1: Collect all raw data
    raw_rows = []
    
    for sheet_data in all_data:
        subject = str(sheet_data.get('subject', sheet_data.get('sheet_name', 'غير محدد'))).strip()
        grade = str(sheet_data.get('grade', '')).strip()
        section = str(sheet_data.get('section', '')).strip()
        
        for student in sheet_data['students']:
            student_name = str(student['student_name']).strip()
            
            # Get values and handle missing/invalid data
            total_assigned = student.get('total_due', 0)
            total_done = student.get('completed', 0)
            
            # Convert to int and clip negative values to 0
            try:
                total_assigned = max(0, int(total_assigned))
            except (ValueError, TypeError):
                total_assigned = 0
            
            try:
                total_done = max(0, int(total_done))
            except (ValueError, TypeError):
                total_done = 0
            
            # Only include if student has assignments (has_due check)
            if student.get('has_due', False):
                raw_rows.append({
                    'student_name': student_name,
                    'grade': grade,
                    'section': section,
                    'subject': subject,
                    'total_assigned': total_assigned,
                    'total_done': total_done
                })
    
    if not raw_rows:
        # Return empty DataFrame with correct columns
        return pd.DataFrame(columns=[
            'student_name',
            'grade',
            'section',
            'subject',
            'subject_total_assigned',
            'subject_total_done',
            'overall_pct_all_subjects',
            'tier'
        ])
    
    # Step 2: Create DataFrame and aggregate by (student, subject)
    df_raw = pd.DataFrame(raw_rows)
    
    # Group by student and subject to aggregate assignments
    df_subject = df_raw.groupby(
        ['student_name', 'grade', 'section', 'subject'],
        as_index=False
    ).agg({
        'total_assigned': 'sum',
        'total_done': 'sum'
    })
    
    # Rename to final column names
    df_subject = df_subject.rename(columns={
        'total_assigned': 'subject_total_assigned',
        'total_done': 'subject_total_done'
    })
    
    # Step 3: Calculate overall percentage for each student across ALL subjects
    df_student_overall = df_subject.groupby(
        ['student_name', 'grade', 'section'],
        as_index=False
    ).agg({
        'subject_total_assigned': 'sum',
        'subject_total_done': 'sum'
    })
    
    # Calculate overall percentage with division by zero protection
    df_student_overall['overall_pct_all_subjects'] = df_student_overall.apply(
        lambda row: round(
            100.0 * row['subject_total_done'] / row['subject_total_assigned']
            if row['subject_total_assigned'] > 0
            else 0.0,
            1
        ),
        axis=1
    )
    
    # Keep only the columns we need for joining
    df_student_overall = df_student_overall[[
        'student_name',
        'grade',
        'section',
        'overall_pct_all_subjects'
    ]]
    
    # Step 4: Join overall percentage back to subject-level data
    df_final = df_subject.merge(
        df_student_overall,
        on=['student_name', 'grade', 'section'],
        how='left'
    )
    
    # Step 5: Calculate tier based on overall percentage
    df_final['tier'] = df_final['overall_pct_all_subjects'].apply(get_tier)
    
    # Step 6: Ensure correct data types
    df_final['subject_total_assigned'] = df_final['subject_total_assigned'].astype(int)
    df_final['subject_total_done'] = df_final['subject_total_done'].astype(int)
    df_final['overall_pct_all_subjects'] = df_final['overall_pct_all_subjects'].astype(float)
    
    # Step 7: Reorder columns to exact specification
    df_final = df_final[[
        'student_name',
        'grade',
        'section',
        'subject',
        'subject_total_assigned',
        'subject_total_done',
        'overall_pct_all_subjects',
        'tier'
    ]]
    
    # Step 8: Sort by grade, section, student_name, subject
    df_final = df_final.sort_values(
        ['grade', 'section', 'student_name', 'subject']
    ).reset_index(drop=True)
    
    return df_final


def export_analytics_to_excel(df: pd.DataFrame, output_path: str) -> str:
    """
    Export analytics DataFrame to Excel with professional formatting.
    
    Args:
        df: Analytics DataFrame
        output_path: Path to save Excel file
    
    Returns:
        str: Path to saved file
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write data
        df.to_excel(writer, sheet_name='Analytics Export', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Analytics Export']
        
        # Format header row
        header_fill = PatternFill(start_color='6d3a46', end_color='6d3a46', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        center_alignment = Alignment(horizontal='center', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Format data cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = right_alignment
                cell.font = Font(name='Arial', size=10)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    return output_path


def export_analytics_to_csv(df: pd.DataFrame, output_path: str) -> str:
    """
    Export analytics DataFrame to CSV.
    
    Args:
        df: Analytics DataFrame
        output_path: Path to save CSV file
    
    Returns:
        str: Path to saved file
    """
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    return output_path

