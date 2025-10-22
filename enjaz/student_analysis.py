"""
Student Analysis Module
Comprehensive student-level analysis with detailed metrics
"""

import pandas as pd
from enjaz.analysis import get_band


def create_student_analysis_table(all_data):
    """
    Create comprehensive student analysis table.
    
    Args:
        all_data: List of sheet data from aggregate_lms_files
    
    Returns:
        pandas.DataFrame with columns:
        - اسم الطالب
        - الصف
        - الشعبة
        - المادة
        - إجمالي المادة
        - المادة منجز
        - المادة متبقي
        - نسبة الحل (%)
        - الفئة
    """
    rows = []
    
    for sheet_data in all_data:
        subject = sheet_data.get('subject', 'غير محدد')
        class_code = sheet_data.get('class_code', 'غير محدد')
        
        # Parse class_code to extract grade and section
        # Format: "03/1" -> Grade: 03, Section: 1
        if '/' in class_code:
            parts = class_code.split('/')
            grade = parts[0].strip()
            section = parts[1].strip() if len(parts) > 1 else 'غير محدد'
        else:
            grade = class_code
            section = 'غير محدد'
        
        for student in sheet_data['students']:
            # Only include students with due assessments
            if not student.get('has_due', False):
                continue
            
            total_due = student.get('total_due', 0)
            completed = student.get('completed', 0)
            not_submitted = student.get('not_submitted', 0)
            remaining = total_due - completed
            completion_rate = student.get('completion_rate', 0.0)
            band = get_band(completion_rate)
            
            rows.append({
                'اسم الطالب': student.get('student_name', 'غير محدد'),
                'الصف': grade,
                'الشعبة': section,
                'المادة': subject,
                'إجمالي المادة': total_due,
                'المادة منجز': completed,
                'المادة متبقي': remaining,
                'نسبة الحل (%)': round(completion_rate, 1),
                'الفئة': band
            })
    
    if not rows:
        return pd.DataFrame()
    
    df = pd.DataFrame(rows)
    
    # Sort by grade, section, subject, student name
    df = df.sort_values(
        by=['الصف', 'الشعبة', 'المادة', 'اسم الطالب'],
        ascending=[True, True, True, True]
    )
    
    return df


def create_student_summary_by_grade(all_data):
    """
    Create summary statistics by grade.
    
    Args:
        all_data: List of sheet data
    
    Returns:
        pandas.DataFrame with grade-level statistics
    """
    df = create_student_analysis_table(all_data)
    
    if df.empty:
        return pd.DataFrame()
    
    summary = df.groupby('الصف').agg({
        'اسم الطالب': 'count',
        'إجمالي المادة': 'sum',
        'المادة منجز': 'sum',
        'نسبة الحل (%)': 'mean'
    }).reset_index()
    
    summary.columns = ['الصف', 'عدد السجلات', 'إجمالي التقييمات', 'المنجز', 'متوسط النسبة (%)']
    summary['متوسط النسبة (%)'] = summary['متوسط النسبة (%)'].round(1)
    
    return summary


def create_student_summary_by_subject(all_data):
    """
    Create summary statistics by subject.
    
    Args:
        all_data: List of sheet data
    
    Returns:
        pandas.DataFrame with subject-level statistics
    """
    df = create_student_analysis_table(all_data)
    
    if df.empty:
        return pd.DataFrame()
    
    summary = df.groupby('المادة').agg({
        'اسم الطالب': 'count',
        'إجمالي المادة': 'sum',
        'المادة منجز': 'sum',
        'نسبة الحل (%)': 'mean'
    }).reset_index()
    
    summary.columns = ['المادة', 'عدد الطلاب', 'إجمالي التقييمات', 'المنجز', 'متوسط النسبة (%)']
    summary['متوسط النسبة (%)'] = summary['متوسط النسبة (%)'].round(1)
    
    return summary


def create_student_summary_by_band(all_data):
    """
    Create summary statistics by performance band.
    
    Args:
        all_data: List of sheet data
    
    Returns:
        pandas.DataFrame with band-level statistics
    """
    df = create_student_analysis_table(all_data)
    
    if df.empty:
        return pd.DataFrame()
    
    summary = df.groupby('الفئة').agg({
        'اسم الطالب': 'count',
        'نسبة الحل (%)': 'mean'
    }).reset_index()
    
    summary.columns = ['الفئة', 'عدد السجلات', 'متوسط النسبة (%)']
    summary['متوسط النسبة (%)'] = summary['متوسط النسبة (%)'].round(1)
    
    # Calculate percentage
    total = summary['عدد السجلات'].sum()
    summary['النسبة من الإجمالي (%)'] = (summary['عدد السجلات'] / total * 100).round(1)
    
    return summary


def export_student_analysis_to_excel(all_data, output_path):
    """
    Export comprehensive student analysis to Excel with multiple sheets.
    
    Args:
        all_data: List of sheet data
        output_path: Path to save Excel file
    
    Returns:
        str: Path to saved file
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Create DataFrames
    df_main = create_student_analysis_table(all_data)
    df_grade = create_student_summary_by_grade(all_data)
    df_subject = create_student_summary_by_subject(all_data)
    df_band = create_student_summary_by_band(all_data)
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_main.to_excel(writer, sheet_name='تحليل الطلاب', index=False)
        df_grade.to_excel(writer, sheet_name='ملخص حسب الصف', index=False)
        df_subject.to_excel(writer, sheet_name='ملخص حسب المادة', index=False)
        df_band.to_excel(writer, sheet_name='ملخص حسب الفئة', index=False)
    
    # Format Excel file
    wb = openpyxl.load_workbook(output_path)
    
    # Define colors
    header_fill = PatternFill(start_color='8A1538', end_color='8A1538', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12, name='Arial')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Center align all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(output_path)
    
    return output_path

