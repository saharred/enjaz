"""
Horizontal Analytics Export Module
Creates detailed analytical DataFrame with one row per student (all subjects in columns)
Matches the exact format of the template file
"""

import pandas as pd
import numpy as np
from typing import List, Dict


def get_tier(overall_pct: float) -> str:
    """
    Get tier classification based on overall percentage.
    
    Args:
        overall_pct: Overall percentage across all subjects
    
    Returns:
        str: Tier name
    """
    if overall_pct >= 90:
        return "بلاتينية"
    elif overall_pct >= 80:
        return "ذهبية"
    elif overall_pct >= 70:
        return "فضية"
    elif overall_pct >= 50:
        return "برونزية"
    elif overall_pct >= 1:
        return "يحتاج إلى تطوير"
    else:  # 0
        return "لا يستفيد من النظام"


def create_horizontal_analytics_export(all_data: List[Dict]) -> pd.DataFrame:
    """
    Create horizontal analytics export DataFrame with one row per student.
    
    Each row contains:
    - الطالب (student name)
    - الصف (grade)
    - الشعبة (section)
    - For each subject (4 columns):
      - [المادة] - إجمالي (total assigned)
      - [المادة] - منجز (completed)
      - [المادة] - النسبة (percentage for this subject)
      - [المادة] - متبقي (remaining)
    - المتوسط (overall average percentage)
    - الفئة (tier)
    
    Args:
        all_data: List of sheet data dictionaries
    
    Returns:
        DataFrame with horizontal analytics export
    """
    
    # Step 1: Collect all student data
    student_data = {}
    all_subjects = set()
    
    for sheet_data in all_data:
        subject = str(sheet_data.get('subject', sheet_data.get('sheet_name', 'غير محدد'))).strip()
        all_subjects.add(subject)
        grade = str(sheet_data.get('grade', '')).strip()
        section = str(sheet_data.get('section', '')).strip()
        
        for student in sheet_data['students']:
            student_name = str(student['student_name']).strip()
            
            if student_name not in student_data:
                student_data[student_name] = {
                    'grade': grade,
                    'section': section,
                    'subjects': {}
                }
            
            # Get values and handle missing/invalid data
            total_assigned = student.get('total_due', 0)
            total_done = student.get('completed', 0)
            
            # Convert to int and clip negative values to 0
            try:
                total_assigned = max(0, int(total_assigned))
            except (ValueError, TypeError):
                total_assigned = 0
            
            try:
                total_done = max(0, int(total_done))
            except (ValueError, TypeError):
                total_done = 0
            
            # Only include if student has assignments
            if student.get('has_due', False):
                # Aggregate if subject already exists for this student
                if subject in student_data[student_name]['subjects']:
                    student_data[student_name]['subjects'][subject]['total_assigned'] += total_assigned
                    student_data[student_name]['subjects'][subject]['total_done'] += total_done
                else:
                    student_data[student_name]['subjects'][subject] = {
                        'total_assigned': total_assigned,
                        'total_done': total_done
                    }
    
    if not student_data:
        # Return empty DataFrame with basic columns
        return pd.DataFrame(columns=['الطالب', 'الصف', 'الشعبة', 'المتوسط', 'الفئة'])
    
    # Step 2: Sort subjects alphabetically
    sorted_subjects = sorted(list(all_subjects))
    
    # Step 3: Create rows for each student
    report_rows = []
    
    for student_name, data in student_data.items():
        if not data['subjects']:
            continue
        
        # Calculate overall percentage
        total_assigned_all = sum(s['total_assigned'] for s in data['subjects'].values())
        total_done_all = sum(s['total_done'] for s in data['subjects'].values())
        overall_pct = round(
            100.0 * total_done_all / total_assigned_all if total_assigned_all > 0 else 0.0,
            1
        )
        
        # Get tier
        tier = get_tier(overall_pct)
        
        # Create row
        row = {
            'الطالب': student_name,
            'الصف': data['grade'],
            'الشعبة': data['section']
        }
        
        # Add columns for each subject (4 columns per subject)
        for subject in sorted_subjects:
            if subject in data['subjects']:
                total_assigned = data['subjects'][subject]['total_assigned']
                total_done = data['subjects'][subject]['total_done']
                subject_pct = round(
                    100.0 * total_done / total_assigned if total_assigned > 0 else 0.0,
                    1
                )
                remaining = max(0, total_assigned - total_done)
            else:
                total_assigned = 0
                total_done = 0
                subject_pct = 0.0
                remaining = 0
            
            row[f'{subject} - إجمالي'] = total_assigned
            row[f'{subject} - منجز'] = total_done
            row[f'{subject} - النسبة'] = subject_pct
            row[f'{subject} - متبقي'] = remaining
        
        # Add overall metrics
        row['المتوسط'] = overall_pct
        row['الفئة'] = tier
        
        report_rows.append(row)
    
    # Step 4: Create DataFrame
    df = pd.DataFrame(report_rows)
    
    # Step 5: Reorder columns to match template
    # الطالب, الصف, الشعبة, [subjects...], المتوسط, الفئة
    base_cols = ['الطالب', 'الصف', 'الشعبة']
    subject_cols = []
    for subject in sorted_subjects:
        subject_cols.extend([
            f'{subject} - إجمالي',
            f'{subject} - منجز',
            f'{subject} - النسبة',
            f'{subject} - متبقي'
        ])
    final_cols = base_cols + subject_cols + ['المتوسط', 'الفئة']
    
    # Reorder
    df = df[final_cols]
    
    # Step 6: Sort by grade, section, student name
    df = df.sort_values(['الصف', 'الشعبة', 'الطالب']).reset_index(drop=True)
    
    # Step 7: Ensure correct data types
    for col in df.columns:
        if ' - إجمالي' in col or ' - منجز' in col or ' - متبقي' in col:
            df[col] = df[col].astype(int)
        elif ' - النسبة' in col or col == 'المتوسط':
            df[col] = df[col].astype(float)
    
    return df


def export_horizontal_analytics_to_excel(df: pd.DataFrame, output_path: str) -> str:
    """
    Export horizontal analytics DataFrame to Excel with professional formatting.
    
    Args:
        df: Horizontal analytics DataFrame
        output_path: Path to save Excel file
    
    Returns:
        str: Path to saved file
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write data
        df.to_excel(writer, sheet_name='التحليل الشامل', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['التحليل الشامل']
        
        # Format header row
        header_fill = PatternFill(start_color='6d3a46', end_color='6d3a46', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Format data cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = right_alignment
                cell.font = Font(name='Arial', size=9)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row and first 3 columns
        worksheet.freeze_panes = 'D2'
    
    return output_path


def export_horizontal_analytics_to_csv(df: pd.DataFrame, output_path: str) -> str:
    """
    Export horizontal analytics DataFrame to CSV.
    
    Args:
        df: Horizontal analytics DataFrame
        output_path: Path to save CSV file
    
    Returns:
        str: Path to saved file
    """
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    return output_path

